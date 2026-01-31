from flask import Flask, render_template, request, jsonify, send_file, Response
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import json
import os
import threading
import re
import csv
from io import StringIO, BytesIO

from dotenv import load_dotenv
load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'google_maps_scraper_2026')

# Global değişkenler - scraping durumu için
scraping_status = {
    'is_running': False,
    'progress': 0,
    'message': '',
    'results': [],
    'total_found': 0,
    'location': '',
    'profession': ''
}

def analyze_phone_number(phone):
    """
    Telefon numarasını analiz eder.
    Türkiye'de cep telefonu numaraları 5 ile başlar (05XX, +905XX, 905XX)
    Sabit hatlar şehir kodlarıyla başlar (0212, 0216, 0312, vb.)
    
    Returns:
        dict: {
            'is_mobile': bool,  # Cep telefonu mu?
            'formatted': str,   # Temizlenmiş numara
            'whatsapp_link': str,  # WhatsApp linki (sadece cep için)
            'display': str  # Görüntüleme formatı
        }
    """
    if not phone:
        return {'is_mobile': False, 'formatted': '', 'whatsapp_link': '', 'display': ''}
    
    # Numarayı temizle (sadece rakamlar)
    cleaned = re.sub(r'[^\d]', '', phone)
    
    # Türkiye numarası formatlarını kontrol et
    is_mobile = False
    whatsapp_number = ''
    
    # Farklı formatları kontrol et
    if cleaned.startswith('90'):
        # +90 veya 0090 formatı
        if len(cleaned) >= 12:
            digit_after_90 = cleaned[2]
            if digit_after_90 == '5':
                is_mobile = True
                whatsapp_number = cleaned[:12]  # 905XXXXXXXXX
    elif cleaned.startswith('0'):
        # 05XX formatı
        if len(cleaned) >= 11:
            digit_after_0 = cleaned[1]
            if digit_after_0 == '5':
                is_mobile = True
                whatsapp_number = '9' + cleaned[:11]  # 0 yerine 90 koy -> 905XXXXXXXXX
    elif cleaned.startswith('5'):
        # 5XX formatı (başında 0 olmadan)
        if len(cleaned) >= 10:
            is_mobile = True
            whatsapp_number = '90' + cleaned[:10]  # 90 ekle -> 905XXXXXXXXX
    
    # WhatsApp linki oluştur
    whatsapp_link = f"https://wa.me/{whatsapp_number}" if is_mobile and whatsapp_number else ''
    
    return {
        'is_mobile': is_mobile,
        'formatted': cleaned,
        'whatsapp_link': whatsapp_link,
        'display': phone
    }

def scrape_google_maps(location, profession, max_results=20):
    """Google Maps'ten dinamik veri çeker - Detaylı bilgilerle"""
    global scraping_status
    
    scraping_status['is_running'] = True
    scraping_status['progress'] = 0
    scraping_status['message'] = 'Tarayıcı başlatılıyor...'
    scraping_status['results'] = []
    scraping_status['location'] = location
    scraping_status['profession'] = profession
    
    # Arama URL'si oluştur
    search_query = f"{location}+{profession}".replace(' ', '+')
    url = f"https://www.google.com/maps/search/{search_query}"
    
    # Chrome ayarları
    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--lang=tr-TR")
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    
    driver = None
    
    try:
        driver = webdriver.Chrome(options=chrome_options)
        scraping_status['message'] = 'Google Maps açılıyor...'
        scraping_status['progress'] = 5
        
        driver.get(url)
        time.sleep(4)
        
        # Çerez popup'ını kabul et
        try:
            accept_button = WebDriverWait(driver, 3).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Kabul')]"))
            )
            accept_button.click()
            time.sleep(1)
        except:
            pass
        
        scraping_status['message'] = 'İşletme listesi yükleniyor...'
        scraping_status['progress'] = 10
        
        # Sonuçların yüklenmesini bekle
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']"))
            )
        except:
            scraping_status['message'] = 'Sonuç bulunamadı!'
            scraping_status['is_running'] = False
            return []
        
        # Önce tüm linkleri topla
        scraping_status['message'] = 'İşletmeler listeleniyor...'
        scrollable_div = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
        
        place_links = []
        last_count = 0
        scroll_attempts = 0
        max_scroll_attempts = 30
        
        while len(place_links) < max_results and scroll_attempts < max_scroll_attempts:
            # Mevcut linkleri topla
            link_elements = driver.find_elements(By.CSS_SELECTOR, "a.hfpxzc")
            
            for elem in link_elements:
                href = elem.get_attribute("href")
                if href and href not in place_links:
                    place_links.append(href)
                    if len(place_links) >= max_results:
                        break
            
            scraping_status['message'] = f'{len(place_links)} işletme bulundu...'
            scraping_status['progress'] = 10 + int((len(place_links) / max_results) * 10)
            
            # Scroll yap
            if len(place_links) < max_results:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div)
                time.sleep(1.5)
                
                # Listenin sonuna ulaşıldı mı?
                try:
                    driver.find_element(By.XPATH, "//*[contains(text(), 'listenin sonuna')]")
                    break
                except:
                    pass
                
                if len(place_links) == last_count:
                    scroll_attempts += 1
                else:
                    scroll_attempts = 0
                    last_count = len(place_links)
        
        # Şimdi her işletmenin detay sayfasına git
        results = []
        total_places = min(len(place_links), max_results)
        
        for i, link in enumerate(place_links[:max_results]):
            try:
                scraping_status['message'] = f'Detaylar alınıyor: {i+1} / {total_places}'
                scraping_status['progress'] = 20 + int((i / total_places) * 75)
                
                # Detay sayfasına git
                driver.get(link)
                time.sleep(2)
                
                # Detayları çek
                result = extract_detailed_data(driver, i + 1, link)
                
                if result and result['isim']:
                    results.append(result)
                    scraping_status['results'] = results.copy()
                    scraping_status['total_found'] = len(results)
                    
            except Exception as e:
                print(f"Hata (işletme {i+1}): {str(e)}")
                continue
        
        scraping_status['progress'] = 100
        scraping_status['message'] = f'Tamamlandı! {len(results)} işletme detayı çekildi.'
        scraping_status['results'] = results
        scraping_status['total_found'] = len(results)
        
        # Sonuçları JSON'a kaydet
        save_results(results, location, profession)
        
        return results
        
    except Exception as e:
        scraping_status['message'] = f'Hata: {str(e)}'
        return []
        
    finally:
        if driver:
            driver.quit()
        scraping_status['is_running'] = False

def extract_detailed_data(driver, index, link):
    """İşletme detay sayfasından tüm bilgileri çeker"""
    result = {
        'sira': index,
        'isim': '',
        'puan': '',
        'degerlendirme_sayisi': '',
        'kategori': '',
        'adres': '',
        'telefon': '',
        'telefon_bilgi': {},  # Telefon analiz bilgileri
        'website': '',
        'calisma_saatleri': {},
        'calisma_durumu': '',
        'plus_code': '',
        'link': link
    }
    
    try:
        # Sayfanın yüklenmesini bekle
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "h1, div.fontHeadlineLarge"))
        )
        time.sleep(1)
        
        # İşletme adı
        try:
            name_selectors = ["h1.DUwDvf", "h1", "div.fontHeadlineLarge"]
            for selector in name_selectors:
                try:
                    name_elem = driver.find_element(By.CSS_SELECTOR, selector)
                    if name_elem.text:
                        result['isim'] = name_elem.text.strip()
                        break
                except:
                    continue
        except:
            pass
        
        # Puan
        try:
            rating_elem = driver.find_element(By.CSS_SELECTOR, "div.F7nice span[aria-hidden='true']")
            result['puan'] = rating_elem.text.strip()
        except:
            try:
                rating_elem = driver.find_element(By.CSS_SELECTOR, "span.ceNzKf")
                result['puan'] = rating_elem.get_attribute("aria-label").split()[0]
            except:
                pass
        
        # Değerlendirme sayısı
        try:
            reviews_elem = driver.find_element(By.CSS_SELECTOR, "div.F7nice span[aria-label*='yorum']")
            reviews_text = reviews_elem.get_attribute("aria-label")
            result['degerlendirme_sayisi'] = re.search(r'([\d.,]+)', reviews_text).group(1) if reviews_text else ''
        except:
            try:
                reviews_elem = driver.find_element(By.CSS_SELECTOR, "span.UY7F9")
                result['degerlendirme_sayisi'] = reviews_elem.text.replace("(", "").replace(")", "").strip()
            except:
                pass
        
        # Kategori
        try:
            cat_elem = driver.find_element(By.CSS_SELECTOR, "button.DkEaL")
            result['kategori'] = cat_elem.text.strip()
        except:
            try:
                cat_elem = driver.find_element(By.CSS_SELECTOR, "span.DkEaL")
                result['kategori'] = cat_elem.text.strip()
            except:
                pass
        
        # Adres - aria-label'dan çek
        try:
            addr_button = driver.find_element(By.CSS_SELECTOR, "button[data-item-id='address']")
            addr_label = addr_button.get_attribute("aria-label")
            if addr_label:
                result['adres'] = addr_label.replace("Adres:", "").strip()
        except:
            try:
                addr_elem = driver.find_element(By.CSS_SELECTOR, "div.Io6YTe.fontBodyMedium.kR99db")
                result['adres'] = addr_elem.text.strip()
            except:
                pass
        
        # Telefon - aria-label'dan çek
        try:
            phone_button = driver.find_element(By.CSS_SELECTOR, "button[data-item-id^='phone']")
            phone_label = phone_button.get_attribute("aria-label")
            if phone_label:
                result['telefon'] = phone_label.replace("Telefon:", "").strip()
        except:
            try:
                phone_buttons = driver.find_elements(By.CSS_SELECTOR, "button.CsEnBe")
                for btn in phone_buttons:
                    aria = btn.get_attribute("aria-label") or ""
                    if "Telefon:" in aria or "telefon" in aria.lower():
                        result['telefon'] = aria.replace("Telefon:", "").strip()
                        break
            except:
                pass
        
        # Telefon numarasını analiz et (cep mi, sabit mi?)
        if result['telefon']:
            result['telefon_bilgi'] = analyze_phone_number(result['telefon'])
        
        # Website
        try:
            web_link = driver.find_element(By.CSS_SELECTOR, "a[data-item-id='authority']")
            result['website'] = web_link.get_attribute("href")
        except:
            try:
                web_buttons = driver.find_elements(By.CSS_SELECTOR, "a.CsEnBe")
                for btn in web_buttons:
                    aria = btn.get_attribute("aria-label") or ""
                    if "web" in aria.lower() or "site" in aria.lower():
                        result['website'] = btn.get_attribute("href")
                        break
            except:
                pass
        
        # Çalışma durumu (Açık/Kapalı)
        try:
            status_elem = driver.find_element(By.CSS_SELECTOR, "span.ZDu9vd span")
            result['calisma_durumu'] = status_elem.text.strip()
        except:
            pass
        
        # Çalışma saatleri tablosu
        try:
            # Çalışma saatleri bölümüne tıkla
            hours_button = driver.find_element(By.CSS_SELECTOR, "div.OqCZI div.OMl5r")
            hours_button.click()
            time.sleep(0.5)
            
            # Tablo satırlarını oku
            rows = driver.find_elements(By.CSS_SELECTOR, "table.eK4R0e tr.y0skZc")
            hours_dict = {}
            for row in rows:
                try:
                    day = row.find_element(By.CSS_SELECTOR, "td.ylH6lf").text.strip()
                    time_text = row.find_element(By.CSS_SELECTOR, "td.mxowUb").text.strip()
                    hours_dict[day] = time_text
                except:
                    continue
            result['calisma_saatleri'] = hours_dict
        except:
            pass
        
        # Plus Code
        try:
            plus_button = driver.find_element(By.CSS_SELECTOR, "button[data-item-id='oloc']")
            plus_label = plus_button.get_attribute("aria-label")
            if plus_label:
                result['plus_code'] = plus_label.replace("Plus code:", "").strip()
        except:
            pass
        
        return result
        
    except Exception as e:
        print(f"Detay çekme hatası: {str(e)}")
        return result

def save_results(results, location, profession):
    """Sonuçları JSON dosyasına kaydet"""
    filename = f"sonuclar_{location}_{profession}.json".replace(' ', '_').lower()
    filepath = os.path.join(os.path.dirname(__file__), filename)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    # Genel sonuçlar dosyasına da kaydet
    general_filepath = os.path.join(os.path.dirname(__file__), 'son_arama.json')
    with open(general_filepath, 'w', encoding='utf-8') as f:
        json.dump({
            'lokasyon': location,
            'meslek': profession,
            'sonuc_sayisi': len(results),
            'sonuclar': results
        }, f, ensure_ascii=False, indent=2)

def load_last_results():
    """Son arama sonuçlarını yükle"""
    filepath = os.path.join(os.path.dirname(__file__), 'son_arama.json')
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return None

@app.route('/')
def index():
    """Ana sayfa"""
    last_search = load_last_results()
    return render_template('index.html', last_search=last_search)

@app.route('/search', methods=['POST'])
def search():
    """Arama başlat"""
    global scraping_status
    
    if scraping_status['is_running']:
        return jsonify({'error': 'Bir arama zaten devam ediyor!'}), 400
    
    location = request.form.get('location', '').strip()
    profession = request.form.get('profession', '').strip()
    max_results = int(request.form.get('max_results', 20))
    
    if not location or not profession:
        return jsonify({'error': 'Lokasyon ve meslek alanları zorunludur!'}), 400
    
    # Arka planda scraping başlat
    thread = threading.Thread(target=scrape_google_maps, args=(location, profession, max_results))
    thread.daemon = True
    thread.start()
    
    return jsonify({'success': True, 'message': 'Arama başlatıldı!'})

@app.route('/status')
def status():
    """Scraping durumunu döner"""
    return jsonify(scraping_status)

@app.route('/results')
def get_results():
    """Sonuçları döner"""
    return jsonify(scraping_status)

@app.route('/export/excel')
def export_excel():
    """Sonuçları Excel (CSV) olarak dışa aktar"""
    try:
        # openpyxl kurulu mu kontrol et, değilse CSV olarak export et
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Son arama sonuçlarını yükle
            last_search = load_last_results()
            if not last_search or not last_search.get('sonuclar'):
                return jsonify({'error': 'Dışa aktarılacak veri bulunamadı!'}), 404
            
            results = last_search['sonuclar']
            location = last_search.get('lokasyon', 'bilinmiyor')
            profession = last_search.get('meslek', 'bilinmiyor')
            
            # Excel workbook oluştur
            wb = Workbook()
            ws = wb.active
            ws.title = "Sonuçlar"
            
            # Başlık stilleri
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Başlıklar
            headers = ['Sıra', 'İşletme Adı', 'Puan', 'Değerlendirme', 'Kategori', 
                      'Adres', 'Telefon', 'Telefon Tipi', 'WhatsApp', 'Website', 
                      'Çalışma Durumu', 'Google Maps Link']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # WhatsApp olan satırlar için yeşil fill
            whatsapp_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
            
            # Verileri yaz
            for row_idx, item in enumerate(results, 2):
                telefon_bilgi = item.get('telefon_bilgi', {})
                is_mobile = telefon_bilgi.get('is_mobile', False)
                whatsapp_link = telefon_bilgi.get('whatsapp_link', '')
                
                row_data = [
                    item.get('sira', ''),
                    item.get('isim', ''),
                    item.get('puan', ''),
                    item.get('degerlendirme_sayisi', ''),
                    item.get('kategori', ''),
                    item.get('adres', ''),
                    item.get('telefon', ''),
                    'Cep' if is_mobile else 'Sabit',
                    whatsapp_link,
                    item.get('website', ''),
                    item.get('calisma_durumu', ''),
                    item.get('link', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    if is_mobile:
                        cell.fill = whatsapp_fill
            
            # Sütun genişliklerini ayarla
            column_widths = [6, 35, 8, 12, 20, 40, 18, 10, 35, 35, 25, 50]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            # BytesIO'ya kaydet
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"sonuclar_{location}_{profession}.xlsx".replace(' ', '_')
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            # openpyxl yoksa CSV olarak export et
            last_search = load_last_results()
            if not last_search or not last_search.get('sonuclar'):
                return jsonify({'error': 'Dışa aktarılacak veri bulunamadı!'}), 404
            
            results = last_search['sonuclar']
            location = last_search.get('lokasyon', 'bilinmiyor')
            profession = last_search.get('meslek', 'bilinmiyor')
            
            output = StringIO()
            writer = csv.writer(output, delimiter=';')
            
            # Başlıklar
            writer.writerow(['Sıra', 'İşletme Adı', 'Puan', 'Değerlendirme', 'Kategori', 
                           'Adres', 'Telefon', 'Telefon Tipi', 'WhatsApp Link', 'Website', 
                           'Çalışma Durumu', 'Google Maps Link'])
            
            # Veriler
            for item in results:
                telefon_bilgi = item.get('telefon_bilgi', {})
                is_mobile = telefon_bilgi.get('is_mobile', False)
                whatsapp_link = telefon_bilgi.get('whatsapp_link', '')
                
                writer.writerow([
                    item.get('sira', ''),
                    item.get('isim', ''),
                    item.get('puan', ''),
                    item.get('degerlendirme_sayisi', ''),
                    item.get('kategori', ''),
                    item.get('adres', ''),
                    item.get('telefon', ''),
                    'Cep' if is_mobile else 'Sabit',
                    whatsapp_link,
                    item.get('website', ''),
                    item.get('calisma_durumu', ''),
                    item.get('link', '')
                ])
            
            output.seek(0)
            filename = f"sonuclar_{location}_{profession}.csv".replace(' ', '_')
            
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    debug = os.getenv('FLASK_ENV', 'production') != 'production'
    port = int(os.getenv('PORT', 5000))
    app.run(debug=debug, host='0.0.0.0', port=port, threaded=True)
